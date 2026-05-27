#!/usr/bin/env python3
"""Build the provisional experiment workbook for Google Sheets import."""

from __future__ import annotations

import json
import os
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPERIMENT_ID = "2026-05-20_taxonomy_augmented_outline_prompt"
DEFAULT_RUN_ID = "2026-05-20_paper096"
PAPER_ID = "096_2502.03108"
SHEET_NAME = "Outline_COT provisional taxonomy augmented outline prompt 2026-05-20"
DRIVE_FOLDER_URL = "https://drive.google.com/drive/folders/1l1bINVVStjrVuhpp6AoS_KPnwpvH0h5W?usp=drive_link"
SCORE_KEYS = [
    "结构_信息快速定位",
    "结构_详略得当",
    "内容_章节互斥性",
    "内容_逻辑深度",
    "内容_学术价值",
    "语用_描述性与简洁性",
]

ROOT_DIR = Path(__file__).resolve().parents[3]
RUN_ID = os.environ.get("TAXONOMY_PROMPT_RUN_ID", DEFAULT_RUN_ID)
RESULTS_ROOT = ROOT_DIR / "results" / "experiments" / EXPERIMENT_ID / RUN_ID
TAXONOMY_PATH = ROOT_DIR / "results" / "experiments" / "2026-05-19_meow_taxonomy_extraction" / "smoke" / PAPER_ID / "taxonomy_extraction.json"
OUTPUT_PATH = RESULTS_ROOT / "_summaries" / f"{SHEET_NAME}.xlsx"
TABLE_PACKAGE_DIR = ROOT_DIR / "_gdrive_sync_outline_cot" / "artifacts" / "tables" / "experiments" / EXPERIMENT_ID
SYNC_SNAPSHOT_PATH = TABLE_PACKAGE_DIR / "snapshots" / f"{SHEET_NAME}.xlsx"


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8") if path.exists() else ""


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def append_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in column_cells[:80]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 55)


def make_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    append_rows(
        ws,
        [
            ["Field", "Value"],
            ["experiment_id", EXPERIMENT_ID],
            ["paper_id", PAPER_ID],
            ["sheet_status", "provisional"],
            ["generated_at_utc", utc_now_iso()],
            ["generation_model", "gpt-5.4-mini"],
            ["generation_reasoning_effort", "medium"],
            ["judge_backend", "codex"],
            ["judge_model", "gpt-5.5"],
            ["judge_reasoning_effort", "high"],
            ["input_conditions", "no_abstract; with_abstract"],
            ["variants", "baseline_no_taxonomy; taxonomy_augmented_v1_minimal; taxonomy_augmented_v2_guarded"],
            ["taxonomy_payload", "tree_only; no status/kind/source boundary/qualifiers/evidence/audit/citation assignments"],
            ["target_drive_folder", DRIVE_FOLDER_URL],
            ["stable_ledger_policy", "Do not merge into the stable ledger until user and Codex explicitly agree."],
        ],
    )
    style_sheet(ws)


def make_run_matrix(wb: Workbook, run_matrix: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Run Matrix")
    headers = [
        "paper_id",
        "input_condition",
        "variant",
        "generation_status",
        "eval_status",
        "model",
        "reasoning_effort",
        "judge_backend",
        "judge_model",
        "judge_reasoning_effort",
        "include_abstract",
        "taxonomy_payload_mode",
        "prompt_character_count",
        "heading_count",
        "max_level",
        "structural_distance",
        "judge_average",
    ]
    ws.append(headers)
    for row in run_matrix:
        stats = row.get("outline_stats", {})
        ws.append(
            [
                row.get("paper_id"),
                row.get("input_condition"),
                row.get("variant"),
                row.get("generation_status"),
                row.get("eval_status"),
                row.get("model"),
                row.get("reasoning_effort"),
                row.get("judge_backend"),
                row.get("judge_model"),
                row.get("judge_reasoning_effort"),
                row.get("include_abstract"),
                row.get("taxonomy_payload_mode"),
                row.get("prompt_character_count"),
                stats.get("heading_count"),
                stats.get("max_level"),
                row.get("structural_distance"),
                row.get("judge_average"),
            ]
        )
    style_sheet(ws)


def make_prompt_variants(wb: Workbook) -> None:
    ws = wb.create_sheet("Prompt Variants")
    ws.append(["variant", "template_path", "description", "key_wording"])
    templates = {
        "baseline_no_taxonomy": (
            "scripts/codex_meow_outline_blind_lib.py",
            "Original blind outline prompt without taxonomy.",
            "Write an outline for a literature review based on the given title and references.",
        ),
        "taxonomy_augmented_v1_minimal": (
            "experiments/2026-05-20_taxonomy_augmented_outline_prompt/prompts/taxonomy_augmented_outline_prompt_template.txt",
            "Minimal taxonomy augmentation.",
            "Use this taxonomy to improve the organization of the main topical/methodological body when it is relevant.",
        ),
        "taxonomy_augmented_v2_guarded": (
            "experiments/2026-05-20_taxonomy_augmented_outline_prompt/prompts/taxonomy_augmented_outline_prompt_guarded_template.txt",
            "Taxonomy augmentation plus explicit anti-overexpansion guidance.",
            "Do not mechanically convert every taxonomy node into a heading. Some taxonomy leaves may belong only as paragraph-level distinctions.",
        ),
    }
    for variant, values in templates.items():
        ws.append([variant, *values])
    style_sheet(ws)


def make_taxonomy_payload(wb: Workbook) -> None:
    ws = wb.create_sheet("Taxonomy Payload")
    ws.append(["field", "value"])
    payload_path = RESULTS_ROOT / PAPER_ID / "no_abstract" / "taxonomy_augmented_v1_minimal" / "taxonomy_tree_payload.txt"
    taxonomy_tree = load_text(payload_path).strip()
    ws.append(["mode", "tree_only"])
    ws.append(["source_artifact", str(TAXONOMY_PATH.relative_to(ROOT_DIR))])
    ws.append(["payload_path", str(payload_path)])
    ws.append(["tree_line_count", len([line for line in taxonomy_tree.splitlines() if line.strip()])])
    ws.append(["tree", taxonomy_tree])
    style_sheet(ws)


def make_generation_results(wb: Workbook, run_matrix: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Generation Results")
    ws.append(["input_condition", "variant", "generation_status", "heading_count", "max_level", "top_level_headings", "taxonomy_leaf_heading_matches", "output_dir"])
    for row in run_matrix:
        stats = row.get("outline_stats", {})
        ws.append(
            [
                row.get("input_condition"),
                row.get("variant"),
                row.get("generation_status"),
                stats.get("heading_count"),
                stats.get("max_level"),
                "; ".join(stats.get("top_level_headings", [])),
                "; ".join(stats.get("taxonomy_leaf_heading_matches", [])),
                row.get("output_dir"),
            ]
        )
    style_sheet(ws)


def make_judge_results(wb: Workbook, run_matrix: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Judge Results")
    ws.append(["input_condition", "variant", "structural_distance", "judge_average", *SCORE_KEYS])
    for row in run_matrix:
        scores = row.get("judge_scores") or {}
        ws.append(
            [
                row.get("input_condition"),
                row.get("variant"),
                row.get("structural_distance"),
                row.get("judge_average"),
                *[scores.get(key) for key in SCORE_KEYS],
            ]
        )
    style_sheet(ws)


def make_pairwise(wb: Workbook, pairwise: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Pairwise Comparison")
    ws.append(
        [
            "input_condition",
            "variant",
            "structural_distance_baseline",
            "structural_distance_variant",
            "structural_distance_improvement_vs_baseline",
            "judge_average_baseline",
            "judge_average_variant",
            "judge_average_delta_vs_baseline",
            "score_deltas_json",
        ]
    )
    for row in pairwise:
        ws.append(
            [
                row.get("input_condition"),
                row.get("variant"),
                row.get("structural_distance_baseline"),
                row.get("structural_distance_variant"),
                row.get("structural_distance_improvement_vs_baseline"),
                row.get("judge_average_baseline"),
                row.get("judge_average_variant"),
                row.get("judge_average_delta_vs_baseline"),
                json.dumps(row.get("score_deltas", {}), ensure_ascii=False),
            ]
        )
    style_sheet(ws)


def make_manual_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("Manual Notes")
    ws.append(["note_type", "note"])
    ws.append(["initial_read", "Minimal taxonomy augmentation improved judge average in both no_abstract and with_abstract conditions."])
    ws.append(["initial_read", "With abstract, minimal taxonomy also improved structural distance versus baseline."])
    ws.append(["initial_read", "Guarded prompt under-generated detail: only 7 or 8 top-level headings and no paragraph-level substructure."])
    ws.append(["manual_check_needed", "Review whether minimal prompt over-expanded taxonomy leaves: both minimal arms include 10 taxonomy leaf labels as headings or near-heading matches."])
    ws.append(["manual_check_needed", "Review whether ordinary scaffold sections remain present. All six arms retain Introduction and Conclusion; most retain Background/Preliminaries."])
    style_sheet(ws)


def main() -> int:
    run_matrix = load_json(RESULTS_ROOT / "_summaries" / "run_matrix.json")
    pairwise = load_json(RESULTS_ROOT / "_summaries" / "paired_comparison.json")
    wb = Workbook()
    make_readme(wb)
    make_run_matrix(wb, run_matrix)
    make_prompt_variants(wb)
    make_taxonomy_payload(wb)
    make_generation_results(wb, run_matrix)
    make_judge_results(wb, run_matrix)
    make_pairwise(wb, pairwise)
    make_manual_notes(wb)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    SYNC_SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUTPUT_PATH, SYNC_SNAPSHOT_PATH)
    print(OUTPUT_PATH)
    print(SYNC_SNAPSHOT_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
